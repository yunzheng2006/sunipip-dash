#!/usr/bin/env python3
"""生成两个 Excel:
1. Spark对账表_20260525.xlsx — 今日余额对账 + 成本分析
2. 上游API全量数据_20260525.xlsx — Spark+IPIPV 全量导出 + 平台对照
"""
import csv, json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
RECONCILE = os.path.join(BASE, "spark-reconcile-20260525")
API_EXPORT = os.path.join(BASE, "api-export-20260525")

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FILL_GREEN = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
HEADER_FILL_ORANGE = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
HEADER_FILL_RED = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
HEADER_FILL_PURPLE = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
HEADER_FILL_TEAL = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
ALT_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def read_csv(path):
    if not os.path.exists(path) or os.path.getsize(path) == 0:
        return []
    with open(path, "r", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def read_json(path):
    if not os.path.exists(path):
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or "")
            width = sum(2 if ord(c) > 127 else 1 for c in val)
            max_len = max(max_len, width)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def style_header(ws, fill=HEADER_FILL):
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_data(ws, start_row=2):
    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row), start=0):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 1:
                cell.fill = ALT_FILL


def write_sheet(ws, rows, fill=HEADER_FILL, number_cols=None):
    if not rows:
        ws.append(["无数据"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for r in rows:
        vals = []
        for h in headers:
            v = r.get(h, "")
            if number_cols and h in number_cols:
                try:
                    v = float(v) if v not in ("", None) else ""
                except (ValueError, TypeError):
                    pass
            vals.append(v)
        ws.append(vals)
    style_header(ws, fill)
    style_data(ws)
    auto_width(ws)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


# Load pricing data
pricing_data = read_json(os.path.join(BASE, "spark_pricing.json"))
COST_MAP = pricing_data.get("cost_map", {})
PRICE_TABLE = pricing_data.get("price_table", [])


def get_spark_cost(product_id):
    info = COST_MAP.get(product_id, {})
    return float(info.get("cost_price", 0))


def get_product_name(product_id):
    info = COST_MAP.get(product_id, {})
    return info.get("product_name", "")


def get_product_country(product_id):
    info = COST_MAP.get(product_id, {})
    return info.get("country_code", "")


# ============================================================
# Excel 1: Spark 对账表 (含成本分析)
# ============================================================
def build_reconcile_excel():
    wb = Workbook()
    all_rows = read_csv(os.path.join(RECONCILE, "spark_orders_reconcile.csv"))

    # 计算 Spark 真实成本
    create_spark = 0
    create_sales = 0
    renew_spark = 0
    renew_sales = 0
    for r in all_rows:
        pid = r.get("产品ID", "")
        cost = get_spark_cost(pid)
        amt = int(r.get("数量", 1) or 1)
        sc = float(r.get("平台成本", 0) or 0)
        method = r.get("方法", "")
        if method == "CreateProxy":
            create_spark += cost * amt
            create_sales += sc * amt if sc else 0
        elif method == "RenewProxy":
            renew_spark += cost
            renew_sales += sc if sc else 0

    # --- Sheet 1: 汇总 ---
    ws = wb.active
    ws.title = "汇总"
    data = [
        ["Spark 余额对账 — 2026-05-25", "", ""],
        [],
        ["项目", "数值", "说明"],
        ["期初余额 (00:11)", 11204, ""],
        ["期末余额 (当前)", 10055, ""],
        ["实际余额消耗", 1149, "11204 - 10055"],
        [],
        ["订单统计", "数量", ""],
    ]
    creates = [r for r in all_rows if r.get("方法") == "CreateProxy"]
    renews = [r for r in all_rows if r.get("方法") == "RenewProxy"]
    deletes = [r for r in all_rows if r.get("方法") == "DelProxy"]
    data += [
        ["新开 (CreateProxy)", len(creates), ""],
        ["续费 (RenewProxy)", len(renews), ""],
        ["删除/退款 (DelProxy)", len(deletes), ""],
        ["合计", len(all_rows), ""],
        [],
        ["成本对比", "平台 sales_cost", "Spark cost_price (实际扣费)"],
        ["新开成本", create_sales, create_spark],
        ["续费成本", renew_sales, renew_spark],
        ["小计", create_sales + renew_sales, create_spark + renew_spark],
        [],
        ["删除退回估算", "", "≈¥62 (志月¥19 + 测试IP¥19 + 香港测试¥24)"],
        ["Spark净消耗估算", "", round(create_spark + renew_spark - 62, 2)],
        ["实际余额消耗", "", 1149],
        ["最终差异", "", round(create_spark + renew_spark - 62 - 1149, 2)],
        [],
        ["结论", "", ""],
        ["① sales_cost ≠ Spark实际扣费", "", "sales_cost 是平台销售底价，每IP比Spark高约¥4"],
        ["② 删除退回减少了实际消耗", "", "释放未用完的IP时Spark按比例退款"],
        ["③ Spark扣费正确", "", "用cost_price计算与实际消耗基本吻合(差¥43)"],
    ]
    for row in data:
        ws.append(row)

    ws["A1"].font = Font(bold=True, size=16, color="4472C4")
    ws.merge_cells("A1:C1")
    for r in [3, 8, 14, 26]:
        for c in range(1, 4):
            ws.cell(row=r, column=c).font = Font(bold=True, size=11)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 38

    # --- Sheet 2: 全部明细 (加入 Spark cost_price) ---
    enriched = []
    for r in all_rows:
        pid = r.get("产品ID", "")
        cost = get_spark_cost(pid)
        amt = int(r.get("数量", 1) or 1)
        sc = float(r.get("平台成本", 0) or 0)
        row = dict(r)
        row["Spark单价"] = cost if cost else ""
        row["Spark总价"] = cost * amt if cost else ""
        row["成本差额"] = round(sc * amt - cost * amt, 2) if cost and sc else ""
        enriched.append(row)

    num_cols = {"数量", "时长", "平台售价", "平台成本", "平台挂牌价", "续费次数",
                "客户ID", "订阅ID", "序号", "Spark单价", "Spark总价", "成本差额"}
    ws2 = wb.create_sheet("全部明细")
    write_sheet(ws2, enriched, HEADER_FILL, num_cols)

    # --- Sheet 3: 新开 ---
    ws3 = wb.create_sheet("新开 CreateProxy")
    write_sheet(ws3, [r for r in enriched if r.get("方法") == "CreateProxy"], HEADER_FILL_GREEN, num_cols)

    # --- Sheet 4: 续费 ---
    ws4 = wb.create_sheet("续费 RenewProxy")
    write_sheet(ws4, [r for r in enriched if r.get("方法") == "RenewProxy"], HEADER_FILL_ORANGE, num_cols)

    # --- Sheet 5: 删除 ---
    ws5 = wb.create_sheet("删除 DelProxy")
    write_sheet(ws5, [r for r in enriched if r.get("方法") == "DelProxy"], HEADER_FILL_RED, num_cols)

    # --- Sheet 6: 成本价对照表 ---
    ws6 = wb.create_sheet("成本价对照")
    country_costs = {}
    for pid, info in COST_MAP.items():
        cc = info.get("country_code", "?")
        cp = float(info.get("cost_price", 0))
        pn = info.get("product_name", "")
        if cc not in country_costs:
            country_costs[cc] = {"products": [], "cost_price": cp}
        country_costs[cc]["products"].append(pn)

    cost_rows = []
    for rule in sorted(PRICE_TABLE, key=lambda x: x.get("priority", 0), reverse=True):
        cc = rule.get("country_code", "*")
        spark_costs = set()
        for pid, info in COST_MAP.items():
            if info.get("country_code") == cc:
                spark_costs.add(float(info.get("cost_price", 0)))
        spark_str = "/".join(str(int(c)) for c in sorted(spark_costs)) if spark_costs else ""
        sales_fixed = rule.get("sales_fixed_price") or ""
        fixed = rule.get("fixed_price") or ""
        diff = ""
        if sales_fixed and spark_costs:
            avg_spark = sum(spark_costs) / len(spark_costs)
            diff = round(float(sales_fixed) - avg_spark, 1)
        cost_rows.append({
            "国家代码": cc,
            "Spark实际扣费(cost_price)": spark_str,
            "平台销售底价(sales_cost)": sales_fixed,
            "平台挂牌价(list_price)": fixed,
            "底价-Spark差额": diff,
            "说明": f"sales_cost比Spark实际高¥{diff}" if isinstance(diff, float) and diff > 0 else "",
        })
    write_sheet(ws6, cost_rows, HEADER_FILL_TEAL)

    out = os.path.join(BASE, "Spark对账表_20260525.xlsx")
    wb.save(out)
    print(f"✓ {out}")


# ============================================================
# Excel 2: 上游API全量 + 平台对照
# ============================================================
def build_api_excel():
    wb = Workbook()
    platform_subs = read_json(os.path.join(BASE, "platform_subscriptions.json"))
    if isinstance(platform_subs, dict):
        platform_subs = []

    sub_by_instance = {}
    for s in platform_subs:
        iid = s.get("spark_instance_id", "")
        if iid:
            sub_by_instance[iid] = s

    # --- Sheet 1: Spark 余额 ---
    ws = wb.active
    ws.title = "Spark余额"
    bal = read_json(os.path.join(API_EXPORT, "spark_balance.json"))
    if bal:
        ws.append(["Spark 账户余额"])
        ws["A1"].font = Font(bold=True, size=14, color="4472C4")
        ws.append([])
        ws.append(["项目", "数值"])
        ws.append(["可用余额", bal.get("amount", "")])
        ws.append(["冻结", bal.get("ext", {}).get("frozen", 0)])
        ws.append(["应收", bal.get("ext", {}).get("receivable", 0)])
        ws.append(["信用额度", bal.get("ext", {}).get("creditLimit", 0)])
        ws.append(["付费模式", bal.get("ext", {}).get("payMode", "")])
        ws.append(["币种", bal.get("ext", {}).get("currency", "")])
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20

    # --- Sheet 2: IPIPV 账户 ---
    ws2 = wb.create_sheet("IPIPV账户")
    ipipv_info = read_json(os.path.join(API_EXPORT, "ipipv_account.json"))
    if ipipv_info:
        ws2.append(["IPIPV 账户信息"])
        ws2["A1"].font = Font(bold=True, size=14, color="548235")
        ws2.append([])
        ws2.append(["项目", "数值"])
        ws2.append(["应用名称", ipipv_info.get("appName", "")])
        ws2.append(["余额", ipipv_info.get("coin", "")])
        ws2.append(["状态", ipipv_info.get("status", "")])
        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 40

    # --- Sheet 3: Spark 实例 + 平台对照 (含 cost_price) ---
    spark_instances = read_csv(os.path.join(API_EXPORT, "spark_instances.csv"))
    merged_rows = []
    for inst in spark_instances:
        iid = inst.get("instance_id", "")
        sub = sub_by_instance.get(iid, {})
        pid = inst.get("product_id", "")
        spark_cost = get_spark_cost(pid)
        sales_cost = sub.get("sales_cost", "")
        try:
            sc_val = float(sales_cost) if sales_cost not in ("", None) else None
        except (ValueError, TypeError):
            sc_val = None
        diff = round(sc_val - spark_cost, 2) if sc_val and spark_cost else ""

        merged_rows.append({
            "instance_id": iid,
            "IP": inst.get("ip", ""),
            "端口": inst.get("port", ""),
            "用户名": inst.get("username", ""),
            "Spark状态": inst.get("status", ""),
            "状态文字": inst.get("status_text", ""),
            "国家": inst.get("country_code", ""),
            "城市": inst.get("city_code", ""),
            "Spark过期": inst.get("expire_time", ""),
            "Spark实际成本": spark_cost if spark_cost else "",
            "产品ID": pid,
            "产品名称": get_product_name(pid),
            "  ": "",
            "平台订阅ID": sub.get("sub_id", ""),
            "客户ID": sub.get("customer_id", ""),
            "客户名称": sub.get("customer_name", ""),
            "平台售价": sub.get("price", ""),
            "管理员定价": sub.get("admin_set_price", ""),
            "挂牌价": sub.get("list_price", ""),
            "平台sales_cost": sales_cost,
            "cost差额": diff,
            "平台状态": sub.get("status", ""),
            "是否测试": sub.get("is_test", ""),
            "自动续费": sub.get("auto_renew", ""),
            "续费次数": sub.get("renewed_count", ""),
            "平台过期": sub.get("expires_at", ""),
            "上游过期": sub.get("upstream_expires_at", ""),
        })

    ws3 = wb.create_sheet("Spark实例+平台对照")
    num_cols3 = {"端口", "Spark状态", "Spark实际成本", "平台订阅ID", "客户ID",
                 "平台售价", "管理员定价", "挂牌价", "平台sales_cost", "cost差额",
                 "是否测试", "自动续费", "续费次数"}
    write_sheet(ws3, merged_rows, HEADER_FILL, num_cols3)

    # --- Sheet 4: Spark 全量订单 ---
    spark_orders = read_csv(os.path.join(API_EXPORT, "spark_orders.csv"))
    ws4 = wb.create_sheet("Spark订单")
    write_sheet(ws4, spark_orders, HEADER_FILL_ORANGE)

    # --- Sheet 5: IPIPV 实例 ---
    ipipv_instances = read_csv(os.path.join(API_EXPORT, "ipipv_instances.csv"))
    ws5 = wb.create_sheet("IPIPV实例")
    write_sheet(ws5, ipipv_instances, HEADER_FILL_GREEN)

    # --- Sheet 6: IPIPV 订单 ---
    ipipv_orders = read_csv(os.path.join(API_EXPORT, "ipipv_orders.csv"))
    ws6 = wb.create_sheet("IPIPV订单")
    write_sheet(ws6, ipipv_orders, HEADER_FILL_GREEN)

    # --- Sheet 7: IPIPV 产品 ---
    ipipv_products = read_csv(os.path.join(API_EXPORT, "ipipv_products.csv"))
    ws7 = wb.create_sheet("IPIPV产品")
    write_sheet(ws7, ipipv_products, HEADER_FILL_GREEN)

    # --- Sheet 8: 平台全量订阅 ---
    ws8 = wb.create_sheet("平台Spark订阅")
    write_sheet(ws8, platform_subs, HEADER_FILL_PURPLE)

    # --- Sheet 9: 差异分析 ---
    diffs = []
    platform_instance_ids = set(s.get("spark_instance_id", "") for s in platform_subs if s.get("spark_instance_id"))
    for inst in spark_instances:
        iid = inst.get("instance_id", "")
        if not iid:
            continue
        spark_status = inst.get("status_text", inst.get("status", ""))
        sub = sub_by_instance.get(iid, {})
        platform_status = sub.get("status", "")

        issue = ""
        if iid not in platform_instance_ids:
            issue = "Spark有实例但平台无订阅"
        elif spark_status in ("released", "releasing") and platform_status == "active":
            issue = "Spark已释放但平台仍active"
        elif spark_status == "active" and platform_status in ("cancelled", "refunded", "expired"):
            issue = "Spark仍active但平台已取消/过期"

        if issue:
            diffs.append({
                "问题": issue,
                "instance_id": iid,
                "IP": inst.get("ip", ""),
                "Spark状态": spark_status,
                "平台状态": platform_status,
                "客户": sub.get("customer_name", ""),
                "订阅ID": sub.get("sub_id", ""),
                "平台售价": sub.get("price", ""),
                "Spark过期": inst.get("expire_time", ""),
                "平台过期": sub.get("expires_at", ""),
            })

    ws9 = wb.create_sheet("差异分析")
    write_sheet(ws9, diffs if diffs else [{"说明": "未发现差异"}], HEADER_FILL_RED)

    # --- Sheet 10: Spark产品价格表 ---
    product_rows = []
    for pid, info in sorted(COST_MAP.items(), key=lambda x: x[1].get("country_code", "")):
        product_rows.append({
            "产品ID": pid,
            "产品名称": info.get("product_name", ""),
            "国家": info.get("country_code", ""),
            "Spark成本价": float(info.get("cost_price", 0)),
        })
    ws10 = wb.create_sheet("Spark产品价格")
    write_sheet(ws10, product_rows, HEADER_FILL_TEAL, {"Spark成本价"})

    out = os.path.join(BASE, "上游API全量数据_20260525.xlsx")
    wb.save(out)
    print(f"✓ {out}")


if __name__ == "__main__":
    build_reconcile_excel()
    build_api_excel()
    print("Done!")
