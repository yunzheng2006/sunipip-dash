#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations
import base64, json, sys, urllib.parse
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    import qrcode
except ImportError:
    print("缺少依赖，请先执行：")
    print("python3 -m pip install openpyxl qrcode[pil] pillow")
    raise SystemExit(1)

DEFAULT_DOMAIN = "hr.sunipip.com"

@dataclass
class NodeInput:
    raw: str
    remark: str
    ip: str
    port: str
    user: str
    password: str

    @property
    def dest(self) -> str:
        return f"{self.ip}:{self.port}"

    @property
    def extract_line(self) -> str:
        return f"{self.remark}##{self.ip}#{self.port}"

@dataclass
class MappingItem:
    dest: str
    listen_port: str
    name: str

def print_title() -> None:
    print("=" * 78)
    print(" Socks5 节点收集与最终表格生成器（支持批量导入） ")
    print("=" * 78)
    print("1. 单条输入：逐条输入 原链接 + 备注")
    print("2. 批量导入：一次性粘贴多行 原链接<TAB>备注")
    print("-" * 78)

def ask_domain() -> str:
    s = input(f"请输入转发域名（直接回车使用默认 {DEFAULT_DOMAIN}）：").strip()
    return s or DEFAULT_DOMAIN

def parse_raw_socks5(raw: str) -> Tuple[str, str, str, str]:
    parts = raw.strip().split(":", 3)
    if len(parts) != 4:
        raise ValueError("格式错误，必须是：IP:端口:用户名:密码")
    ip, port, user, password = [p.strip() for p in parts]
    if not all([ip, port, user, password]):
        raise ValueError("链接中存在空字段")
    return ip, port, user, password

def build_node(raw: str, remark: str) -> NodeInput:
    ip, port, user, password = parse_raw_socks5(raw)
    return NodeInput(raw=raw.strip(), remark=remark.strip(), ip=ip, port=port, user=user, password=password)

def collect_nodes_single() -> List[NodeInput]:
    nodes = []
    while True:
        raw = input("\n请输入 socks5 原链接（IP:端口:用户名:密码）：").strip()
        if not raw:
            print("原链接不能为空。")
            continue
        remark = input("请输入对应备注：").strip()
        if not remark:
            print("备注不能为空。")
            continue
        try:
            node = build_node(raw, remark)
        except ValueError as exc:
            print(f"输入有误：{exc}")
            continue
        nodes.append(node)
        print(f"已添加：{node.remark} -> {node.dest}")
        while True:
            more = input("是否继续添加？(y/n)：").strip().lower()
            if more in {"y", "yes"}:
                break
            if more in {"n", "no"}:
                return nodes
            print("请输入 y 或 n。")

def read_multiline(label: str) -> str:
    print(label)
    print("结束方式：")
    print(" - 粘贴完后按 Ctrl+D")
    print(" - 或单独输入一行 END 再回车")
    lines = []
    try:
        while True:
            line = input()
            if line.strip().upper() == "END":
                break
            lines.append(line)
    except EOFError:
        pass
    return "\n".join(lines).strip()

def collect_nodes_batch() -> Tuple[List[NodeInput], List[str]]:
    text = read_multiline(
        "\n请批量粘贴节点，每行格式必须是：\n原链接<TAB>备注\n\n例如：\n185.78.169.60:41464:5A2v8fA1857816960A41464:lgR32XNZ9g4X\t德国-185.78.169.60\n"
    )
    nodes, bad_lines = [], []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if "\t" in line:
            raw, remark = line.split("\t", 1)
        else:
            parts = line.rsplit(None, 1)
            if len(parts) != 2:
                bad_lines.append(raw_line)
                continue
            raw, remark = parts[0], parts[1]
        try:
            nodes.append(build_node(raw, remark))
        except Exception:
            bad_lines.append(raw_line)
    return nodes, bad_lines

def choose_input_mode() -> str:
    while True:
        print("\n请选择录入方式：\n1. 单条输入\n2. 批量导入")
        choice = input("输入 1 或 2：").strip()
        if choice in {"1", "2"}:
            return choice
        print("请输入 1 或 2。")

def write_extract_file(nodes: List[NodeInput]) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = Path.cwd() / f"{ts}_提取结果.txt"
    path.write_text("\n".join(node.extract_line for node in nodes), encoding="utf-8")
    return path

def show_extract_result(nodes: List[NodeInput]) -> None:
    print("\n" + "=" * 78)
    print("请复制下面这些内容，拿去面板转换：")
    print("=" * 78)
    for node in nodes:
        print(node.extract_line)
    print("=" * 78)
    print('面板返回示例：{"dest":["185.78.169.60:41464"],"listen_port":39999,"name":"德国-185.78.169.60"}')
    print("=" * 78 + "\n")

def parse_mapping_text(text: str) -> Tuple[Dict[str, List[MappingItem]], List[str]]:
    mapping = defaultdict(list)
    bad_lines = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        try:
            obj = json.loads(line)
            dest_list = obj.get("dest")
            listen_port = obj.get("listen_port")
            name = obj.get("name")
            if not isinstance(dest_list, list) or not dest_list or not isinstance(dest_list[0], str):
                raise ValueError("dest 字段错误")
            if listen_port is None or not isinstance(name, str):
                raise ValueError("字段缺失")
            item = MappingItem(dest=dest_list[0].strip(), listen_port=str(listen_port).strip(), name=name.strip())
            mapping[item.dest].append(item)
        except Exception:
            bad_lines.append(raw_line)
    return mapping, bad_lines

def choose_mapping(node: NodeInput, candidates: List[MappingItem]) -> MappingItem:
    exact = [c for c in candidates if c.name == node.remark]
    if exact:
        return exact[-1]
    if len(candidates) == 1:
        return candidates[0]
    return candidates[-1]

def build_v2ray_socks_uri(domain: str, listen_port: str, user: str, password: str, remark: str) -> str:
    auth_b64 = base64.b64encode(f"{user}:{password}".encode("utf-8")).decode("utf-8")
    return f"socks://{urllib.parse.quote(auth_b64, safe='')}@{domain}:{listen_port}#{urllib.parse.quote(remark, safe='')}"

def build_output_rows(nodes, mapping, domain):
    matched_rows, unmatched_nodes, duplicate_cases = [], [], []
    for node in nodes:
        candidates = mapping.get(node.dest, [])
        if not candidates:
            unmatched_nodes.append(node)
            continue
        if len(candidates) > 1:
            duplicate_cases.append((node, candidates))
        chosen = choose_mapping(node, candidates)
        matched_rows.append({
            "socks5原链接": node.raw,
            "socks5已转发": f"{domain}:{chosen.listen_port}:{node.user}:{node.password}",
            "socks5新链接": build_v2ray_socks_uri(domain, chosen.listen_port, node.user, node.password, chosen.name),
            "备注": chosen.name,
        })
    return matched_rows, unmatched_nodes, duplicate_cases

def style_sheet(ws):
    headers = ["socks5原链接", "socks5已转发", "socks5新链接", "备注", "二维码"]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E2F3")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = {1: 42, 2: 46, 3: 78, 4: 28, 5: 16}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = Border(bottom=thin_gray)
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{ws.max_row}"

def create_excel(rows: List[dict], output_dir: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = output_dir / f"{timestamp}_最终结果.xlsx"
    qr_dir = output_dir / f"{timestamp}_二维码缓存"
    qr_dir.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "最终结果"
    ws.append(["socks5原链接", "socks5已转发", "socks5新链接", "备注", "二维码"])
    for idx, row in enumerate(rows, start=2):
        ws.cell(idx, 1, row["socks5原链接"])
        ws.cell(idx, 2, row["socks5已转发"])
        ws.cell(idx, 3, row["socks5新链接"])
        ws.cell(idx, 4, row["备注"])
        for col in range(1, 5):
            ws.cell(idx, col).alignment = Alignment(vertical="center", wrap_text=True)
        qr = qrcode.QRCode(box_size=3, border=2)
        qr.add_data(row["socks5新链接"])
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        img_path = qr_dir / f"qr_{idx}.png"
        img.save(img_path)
        xl_img = XLImage(str(img_path))
        xl_img.width = 96
        xl_img.height = 96
        ws.add_image(xl_img, f"E{idx}")
        ws.row_dimensions[idx].height = 78
    style_sheet(ws)
    wb.save(out_path)
    return out_path

def write_log(output_dir, nodes, bad_input_lines, bad_json_lines, unmatched_nodes, duplicate_cases):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = output_dir / f"{ts}_处理说明.txt"
    lines = [
        f"输入节点数：{len(nodes)}",
        f"批量导入失败行数：{len(bad_input_lines)}",
        f"未解析 JSON 行数：{len(bad_json_lines)}",
        f"未匹配节点数：{len(unmatched_nodes)}",
        f"映射重复节点数：{len(duplicate_cases)}",
        ""
    ]
    if bad_input_lines:
        lines += ["【批量导入失败行】", *bad_input_lines, ""]
    if bad_json_lines:
        lines += ["【未解析的 JSON 行】", *bad_json_lines, ""]
    if unmatched_nodes:
        lines += ["【未匹配节点】", *[f"{n.remark} | {n.raw}" for n in unmatched_nodes], ""]
    if duplicate_cases:
        lines.append("【存在多个 mapping 候选，程序已优先按备注匹配，否则取最后一条】")
        for node, candidates in duplicate_cases:
            lines.append(f"- 节点：{node.remark} | {node.dest}")
            for c in candidates:
                lines.append(f"    候选：listen_port={c.listen_port}, name={c.name}")
        lines.append("")
    if not (bad_input_lines or bad_json_lines or unmatched_nodes or duplicate_cases):
        lines.append("本次处理没有异常。")
    log_path.write_text("\n".join(lines), encoding="utf-8")
    return log_path

def main():
    print_title()
    domain = ask_domain()
    bad_input_lines = []
    mode = choose_input_mode()
    if mode == "1":
        nodes = collect_nodes_single()
    else:
        nodes, bad_input_lines = collect_nodes_batch()
        print(f"\n批量导入完成：成功 {len(nodes)} 条，失败 {len(bad_input_lines)} 条。")
    if not nodes:
        print("没有可用节点，程序结束。")
        return
    extract_file = write_extract_file(nodes)
    show_extract_result(nodes)
    print(f"提取结果已保存到：{extract_file}")
    mapping_text = read_multiline("请粘贴面板返回的 JSON 结果：")
    if not mapping_text:
        print("没有接收到 mapping 结果，程序结束。")
        return
    mapping, bad_json_lines = parse_mapping_text(mapping_text)
    rows, unmatched_nodes, duplicate_cases = build_output_rows(nodes, mapping, domain)
    if not rows:
        print("没有匹配成功的节点，未生成表格。")
        return
    output_dir = Path.cwd()
    excel_path = create_excel(rows, output_dir)
    log_path = write_log(output_dir, nodes, bad_input_lines, bad_json_lines, unmatched_nodes, duplicate_cases)
    print("\n" + "=" * 78)
    print("处理完成。")
    print(f"成功匹配：{len(rows)} / {len(nodes)}")
    print(f"最终表格：{excel_path}")
    print(f"处理说明：{log_path}")
    print("=" * 78)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n已中断。")
        sys.exit(1)
